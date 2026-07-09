from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


START_ROW = 6


work_file = input("Введите название файла на обработку: ")
db_file = input("Введите название файла с БД: ")

wb_wf = load_workbook(work_file)
ws_wf = wb_wf.active
wf_max_col = get_column_letter(ws_wf.max_column + 1)
wf_max_row = ws_wf.max_row

wb_bd = load_workbook(db_file)
ws_db = wb_bd.active

for num_r, r in enumerate(ws_wf, 1):
    if num_r < START_ROW:
        continue
    if r[0].value:
        sku_name = r[0].value.strip()
        for i in ws_db:
            if i[0].value:
                if sku_name == i[0].value.strip():
                    cell = f"{wf_max_col}{num_r}"
                    ws_wf[cell] = i[1].value

wb_wf.save(work_file)

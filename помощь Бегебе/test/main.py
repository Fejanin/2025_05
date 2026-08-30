import excel2img
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from test import combin_png


file = "test.xlsx"
wb = load_workbook(file)
ws = wb.active

max_col_abc = get_column_letter(ws.max_column)
shops = {"head": "Лист1!A1:O1",}
city_num = {"Магазин номер": None, "Город": None}

for i in ws:
    for n, col in enumerate(i, 1):
        if col.value in city_num:
            city_num[col.value] = n
    break

id = 0

for num_row, row in enumerate(ws, 1):
    if num_row == 1:
        continue
    city = row[city_num["Город"] - 1].value
    num_shop = row[city_num["Магазин номер"] - 1].value
    if city and num_shop:
        new_name = f"{city} №{num_shop}({id})"  # ДОЛЖНЫ БЫТЬ ОДИНАКОВЫМИ
        test_name = new_name
        if new_name in shops:
            shops[new_name] = shops[new_name].split(":")[0] + f":{max_col_abc}{num_row}"
        else:
            new_name = f"{city} №{num_shop}({id})"  # ДОЛЖНЫ БЫТЬ ОДИНАКОВЫМИ
            shops[new_name] = f"Лист1!A{num_row}:{max_col_abc}{num_row}"
    elif (city is None) + (num_shop is None) == 1:
        raise ValueError(f"В строке {num_row} отсутствует номер магазина или не указан город.")
    else:
        id += 1

all_files = []

for t in shops:
    tmp_file = f"tmp_data\\{t}.png"
    excel2img.export_img(file, tmp_file, "", shops[t])
    all_files.append(tmp_file)

for img_data in all_files:
    if "head" in img_data:
        img_head = img_data
        continue
    combin_png(img_head, img_data, img_data.split("\\")[1][:-4])
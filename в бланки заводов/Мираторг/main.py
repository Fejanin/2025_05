import openpyxl

base_order = {}
res_order = {}
errors = []
control_matrix = []
control_blank = []
report = []
total = 0
total_res = 0
COL = "O"  # колонка в кот-ую вносим заказ

file = input("Введите название файла с расчетом: ")
file_res = input("Введите название файла бланка завода: ")
try:
    matrix_file = "matrix.xlsx"

    wb = openpyxl.load_workbook(file)
    ws = wb.active

    for r in ws:
        line = [i.value for i in r]
        if line[1]:
            base_order[line[0].strip()] = line[1]

    wb_m = openpyxl.load_workbook(matrix_file)
    ws_m = wb_m.active

    for r in ws_m:
        art = r[0].value.strip()
        if art in base_order:
            res_order[r[1].value] = [art, base_order[art]]
            control_matrix.append(art)
    for i in base_order:
        if i not in control_matrix:
            errors.append(f"В matrix.xlsx отсутствует: {i}.")

    wb_res = openpyxl.load_workbook(file_res)
    ws_res = wb_res.active

    for r in ws_res:
        art = r[0].value
        if art in res_order:
            num_row = str(r[0])[:-1].split(".")[-1][1:]
            val = ws_res[f"{COL}{num_row}"].value
            if val:
                for i in res_order.items():
                    if art in i:
                        res_order[art].append(f"{COL}{num_row} содержит значение - {val}")
                        errors.append(f"{art}: {res_order[art]}.")
            else:
                ws_res[f"{COL}{num_row}"].value = res_order[art][1]
                control_blank.append(art)
                report.append(f"Добавлено {art} => {res_order[art]}, в количестве: {res_order[art][1]} кор.")
                total_res += res_order[art][1]

    for i in res_order:
        if i not in control_blank and res_order[i] == 2:
            errors.append(f"В бланке завода отсутствует - {i}: {res_order[i]}.")

    total = sum(base_order.values())

    with open("REPORT.txt", 'w') as f:
        if errors:
            f.write("ВЫЯВЛЕНЫ ОШИБКИ:\n\n")
            for l in errors:
                f.write(f"{l}\n")
            f.write(f"\n{'#' * 50}\n")
        for i in report:
            f.write(f"{i}\n")
        f.write(f"\n{'#' * 50}\n")
        f.write(f"\n\nКоличество коробок в заказе: {total}\n")
        f.write(f"Количество коробок перенесенных в бланк: {total_res}")

    wb_res.save(file_res)
except Exception as e:
    print(e)
    input()

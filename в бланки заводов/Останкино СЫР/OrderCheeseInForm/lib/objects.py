if __name__ != "__main__":
    from openpyxl import load_workbook
    from lib.funcs import get_columns_abc_by_name, get_max_col, get_max_row
    from lib.DB import db
else:
    from openpyxl import load_workbook
    from funcs import get_columns_abc_by_name, get_max_col, get_max_row
    from DB import db


class Division:
    HEADER = 3
    FIRST_ROW = 6
    
    def __new__(cls, *args, **kwargs):
        if args[0]:
            print("Заказ будет.", args[0])
            return super().__new__(cls)
        else:
            print("Заказа не будет.", args[0])

    def __init__(self, file, name):
        self.file = file
        self.name = name
        self.orders = []
        self.not_in_db = {}
        self.errors = {}
        self.wb_data_only = self.create_wb(data_only=True)
        self.sheet_data_only = self.wb_data_only["Sheet"]
        self.name_cols = [
            "Номенклатура",
            "итого"]
        self.cols_abc = get_columns_abc_by_name(self.name_cols, self.HEADER, self.sheet_data_only)
        self.max_col = get_max_col(self.HEADER, self.sheet_data_only)
        self.max_row = get_max_row(self.sheet_data_only)
        self.get_order_from_file()
        self.weight = 0  # вес итогового заказа

    def create_wb(self, data_only=False):
        return load_workbook(self.file, data_only=data_only)

    def get_order_from_file(self):
        for r in range(self.FIRST_ROW, self.max_row):
            sku = self.sheet_data_only[f"{self.cols_abc['Номенклатура']}{r}"].value
            order = self.sheet_data_only[f"{self.cols_abc['итого']}{r}"].value
            if order:
                if sku in db:
                    self.orders.append(SKU(sku, order))
                else:
                    self.not_in_db[sku] = order


class SKU:
    def __init__(self, name, order):
        self.name = name  # str
        self.order = order  # int or float
        self.article = db[self.name]['art']  # str
        self.name_name = db[self.name]['name']  # str
        self.multiplicity = db[self.name]['mult']  # int or float
        self.type_formuls = db[self.name]['formuls']  # str | 'штуки' / 'вес'
        self.coof = db[self.name]['coof']  # int or float
        self.order_to_factory = self.calculate_order()
        self.weight = self.calculate_weight()
        self.in_order = False

    def calculate_order(self):
        res = round(self.order / self.multiplicity) * self.multiplicity
        return res or self.multiplicity

    def calculate_weight(self):
        return self.coof * self.order_to_factory


'''
TODO
- перенести заказы в бланк
- отследить все ошибки
    - скю нет вланке завода
    - есть отличия в данных SKU и в данных бланка
    - попытка повторного занесения значения в ячейку
- создание отчета по перенесенным данным
- создание отчета по ошибкам
- создание файлов с результирующими заказами, для переноса с помощью ВПР
'''
class ORDER:
    SHEETS = ["Мелитополь", "Бердянск", "Донецк", "Луганск"]
    ARTICLE = "A"
    NAME = "B"
    MULT = "C"  # кратность, шт (только для штучной продукции)
    ORDER_PIECE_PRODUCTS = "D"  # шучная продукция (колонка для заказа)
    ORDER_WEIGHT_PRODUCTS = "E"  # весовая продукция (колонка для заказа)
    PIECE_OR_WEIGHT = "F"  # =D4/C4 ==> штучная; =E47/5 ==> весовая
    COOF = "G"  # вес за единицу (колонка для заказа)
    WEIGHT = "H"
    FIRST_ROW = 4
    
    def __init__(self, file_name, divisions):
        self.file_name = file_name
        self.divisions = divisions
        self.wb = load_workbook(self.file_name)
        self.sheets = self.wb.sheetnames
        if set(self.sheets) != set(self.SHEETS):
            raise Exception("В бланке для заказов названия для страниц не соответствует заданным требованиям!")
        self.create_orders()
        self.wb.save('заказ СЫРЫ от  на .xlsx')

    def create_orders(self):
        for obj in self.divisions:
            self.sheet = self.wb[obj.name]
            self.max_row = get_max_row(self.sheet)
            self.create_order_for_division(obj)

    def create_order_for_division(self, division):
        articles = {sku.article: sku for sku in division.orders}
        for row in range(self.FIRST_ROW, self.max_row + 1):
            sku = str(self.sheet[f'{self.ARTICLE}{row}'].value)
            if sku in articles:
                res = self.check_data(row, articles[sku])
                if res:  # если ошибка
                    if not sku in division.errors:
                        division.errors[sku] = [(division.name, articles[sku].article, articles[sku].name, '\nРасхождение данных между БД и бланком.\n\n')]
                    else:
                        division.errors[sku].append((division.name, articles[sku].article, articles[sku].name, '\nРасхождение данных между БД и бланком.\n\n'))
                    continue
                if self.sheet[f'{self.ORDER_PIECE_PRODUCTS}{row}'].value or self.sheet[f'{self.ORDER_WEIGHT_PRODUCTS}{row}'].value:  # если в ячейке уже есть значение
                    if not sku in division.errors:
                        division.errors[sku] = [(division.name, articles[sku].article, articles[sku].name, '\nВ бланке заказа уже есть значение.\n\n')]
                    else:
                        division.errors[sku].append((division.name, articles[sku].article, articles[sku].name, '\nВ бланке заказа уже есть значение.\n\n'))
                    continue
                # если нет ошибки
                if articles[sku].type_formuls == 'штуки':
                    self.sheet[f'{self.ORDER_PIECE_PRODUCTS}{row}'].value = articles[sku].order_to_factory
                    division.weight += articles[sku].weight
                    articles[sku].in_order = True
                elif articles[sku].type_formuls == 'вес':
                    self.sheet[f'{self.ORDER_WEIGHT_PRODUCTS}{row}'].value = articles[sku].order_to_factory
                    division.weight += articles[sku].weight
                    articles[sku].in_order = True

    def check_data(self, row, sku):
        # проверка данных по БД
        if self.sheet[f'{self.NAME}{row}'].value != sku.name_name:
            return True
        if sku.type_formuls == 'штуки':
            if self.sheet[f'{self.MULT}{row}'].value != sku.multiplicity or \
               self.sheet[f'{self.PIECE_OR_WEIGHT}{row}'].value.split('/')[1] != f'{self.MULT}{row}' or \
               self.sheet[f'{self.COOF}{row}'].value != sku.coof or \
               (self.sheet[f'{self.WEIGHT}{row}'].value != f'={self.COOF}{row}*{self.ORDER_PIECE_PRODUCTS}{row}' and \
                self.sheet[f'{self.WEIGHT}{row}'].value != f'={self.ORDER_PIECE_PRODUCTS}{row}*{self.COOF}{row}'):
                return True
        elif sku.type_formuls == 'вес':
            if self.sheet[f'{self.PIECE_OR_WEIGHT}{row}'].value.split('/')[1] != str(sku.multiplicity) or \
               self.sheet[f'{self.WEIGHT}{row}'].value != f'={self.ORDER_WEIGHT_PRODUCTS}{row}':
                return True
        else:
            raise Exception(f"В базе данных не указано значение вес/штука для {sku.name}({sku.article}).")
        return False


class REPORT:
    def __init__(self, divisions):
        self.divisions = divisions
        self.create_report()
        self.create_error_report()

    def create_report(self):
        '''
        Создаем отчет и заказники в формате txt.
        В отчете указываем какие скю перенесены в бланк заказа, в каком количестве и полученный вес в разрезе филиалов.
        В заказниках по-филиально указываем название скю (из 1С) и заказанное количество (через табуляцию)
        '''
        for d in self.divisions:
            with open(f'заказ {d.name}.txt', 'w') as f:
                for sku in d.orders:
                    f.write(f'{sku.name}\t{round(sku.order_to_factory)}\n')
        with open('REPORT.txt', 'w') as f:
            for d in self.divisions:
                f.write(f'\n{d.name}\n')
                for s in d.orders:
                    f.write(f'({s.article}): {s.name} ==> {round(s.order_to_factory)}\n\t{s.name_name}\n')

    def create_error_report(self):
        '''
        Создаем единый отчет в котором поочередно указываем:
            - скю, которые не перенесены в бланк по причине их отсутствия в БД (новинки, масло, ...)
            - скю по которым были выявлены ошибки.
        '''
        with open('ERRORS.txt', 'w') as f:
            for d in self.divisions:
                if d.not_in_db:
                    f.write(f'{d.name}\nОтсутствуют в БД:\n')
                    for name in d.not_in_db:
                        f.write(f'{name}: {d.not_in_db[name]}\n')
                    f.write('\n')
            for d in self.divisions:
                if d.errors:
                    f.write('#################\n')
                    f.write(f'{d.name}\nВыявлены следующие ошибки:\n')
                    for error in d.errors:
                        f.write(f'{error}: {d.errors[error]}\n')
                    
                    


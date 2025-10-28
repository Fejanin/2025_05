from openpyxl import load_workbook
from openpyxl.utils import FORMULAE


SU = "A"
SU_weight_brutto = "Y"
WEIGHT_NETTO_box = "H"
WEIGHT_BRUTTO_box = "I"
NUM_BOXES_PER_PALLET = "J"

RESULT_BRUTTO = "P"
SUM_PALLET = "P"
TOTAL_WEIGHT_WITH_PALLET = "P"


# file = "test.xlsx"
file = input("Введите название бланка для завода: ")
wb = load_workbook(file)
sh = wb.active
calc_weight = "="
calc_pallets = "ОКРУГЛВВЕРХ("


for row in range(1, sh.max_row + 1):
    if type(sh[f"{SU}{row}"].value) is str and "SU" in sh[f"{SU}{row}"].value:
        calc_weight += f"{SU_weight_brutto}{row}/{WEIGHT_NETTO_box}{row}*{WEIGHT_BRUTTO_box}{row}+"
        calc_pallets += f"{SU_weight_brutto}{row}/{WEIGHT_NETTO_box}{row}/{NUM_BOXES_PER_PALLET}{row}+"
    if type(sh[f"{RESULT_BRUTTO}{row}"].value) is str and sh[f"{RESULT_BRUTTO}{row}"].value == "ИТОГО БРУТТО":
        calc_weight = calc_weight[:-1]
        sh[f"{SU_weight_brutto}{row}"] = calc_weight
        weight_brutto = f"{SU_weight_brutto}{row}"
    if type(sh[f"{SUM_PALLET}{row}"].value) is str and sh[f"{SUM_PALLET}{row}"].value == "Кол-во паллет":
        calc_pallets = calc_pallets[:-1] + ";0)"
        sh[f"{SU_weight_brutto}{row}"] = calc_pallets
        count_pallet = f"{SU_weight_brutto}{row}"
    if type(sh[f"{TOTAL_WEIGHT_WITH_PALLET}{row}"].value) is str and sh[f"{TOTAL_WEIGHT_WITH_PALLET}{row}"].value == "Вес брутто  с паллетами":
        sh[f"{SU_weight_brutto}{row}"] = f"={weight_brutto}" + f"+{count_pallet}*25"

wb.save(f"{file}")
